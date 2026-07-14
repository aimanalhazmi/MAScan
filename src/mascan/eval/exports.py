"""Human-friendly exports for gold-standard prompts and calibration packets."""

import csv
import hashlib
import io
import json
from collections.abc import Sequence
from pathlib import Path

from mascan.eval.fingerprints import model_sha256
from mascan.eval.gold_judge import GOLD_JUDGE_SYSTEM_PROMPT
from mascan.eval.gold_judge import build_gold_judge_user_prompt
from mascan.eval.gold_judge import gold_judge_output_schema
from mascan.eval.gold_judge import gold_judge_prompt_sha256
from mascan.eval.gold_judge import gold_judge_schema_sha256
from mascan.eval.gold_standard import GoldStandardDataset, PESTEL_HEADINGS
from mascan.eval.gold_standard import EXPECTED_OUTPUT_FIELDS
from mascan.eval.gold_standard import GoldStandardCoverageReport
from mascan.eval.gold_standard import GoldStandardCase
from mascan.eval.gold_standard import validate_gold_standard_coverage
from mascan.eval.human_calibration import HumanCalibrationPacket
from mascan.eval.human_ratings import HumanRatingsTemplate


PROMPT_PACK_CSV_FIELDS = ["case_id", "case_title", "source_pdf", "prompt"]
GOLD_STANDARD_MANIFEST_CSV_FIELDS = [
    "case_id",
    "case_title",
    "source_pdf",
    "prompt_sha256",
    "expected_output_sha256",
    "gold_claims_sha256",
    "category_targets_sha256",
    "gold_claim_count",
    "category_target_count",
    "source_anchor_count",
    "expected_sections_complete",
    "category_targets_cover_all_buckets",
]
RATINGS_CSV_FIELDS = [
    "metric",
    "rater_id",
    "case_id",
    "label",
    "factor",
    "correct_category",
    "rationale",
    "analytical_depth_score",
    "correct",
]

def render_prompt_pack_markdown(dataset: GoldStandardDataset) -> str:
    lines = [
        "# Gold-Standard PESTEL Prompt Pack",
        "",
        "Use the same prompt for every control group.",
    ]
    for case in dataset.cases:
        lines += [
            "",
            "---",
            "",
            f"## {case.case_id}: {case.case_title}",
            "",
            f"**Source PDF:** `{case.source_pdf}`",
            "",
            "```text",
            case.prompt,
            "```",
        ]
    return "\n".join(lines) + "\n"


def render_gold_standard_validation_report(
    dataset: GoldStandardDataset,
    coverage_report: GoldStandardCoverageReport | None = None,
) -> str:
    """Render prompt, expected answer, and reread justification for each case."""
    coverage_report = coverage_report or validate_gold_standard_coverage(dataset)
    all_expected_sections_present = all(
        all(getattr(case.expected_output, field) for field in EXPECTED_OUTPUT_FIELDS)
        for case in dataset.cases
    )
    all_target_buckets_present = all(
        {target.correct_category for target in case.category_targets}
        == set(PESTEL_HEADINGS)
        for case in dataset.cases
    )
    target_counts = [len(case.category_targets) for case in dataset.cases]
    claim_counts = [len(case.gold_claims) for case in dataset.cases]

    lines = [
        "# Gold-Standard Dataset Validation Report",
        "",
        "This report exposes the prompt, expected PESTEL answer, source anchors, "
        "and reread justification for every case in the frozen dataset.",
        "",
        "## Consistency Summary",
        "",
        f"- Cases: {len(dataset.cases)}",
        f"- PDF papers in inventory: {coverage_report.pdf_count}",
        f"- Dataset/PDF coverage exact: {_yes_no(coverage_report.is_valid)}",
        f"- Expected output sections complete: {_yes_no(all_expected_sections_present)}",
        f"- Categorization targets cover all PESTEL buckets per case: "
        f"{_yes_no(all_target_buckets_present)}",
        f"- Categorization targets per case: {min(target_counts)}-{max(target_counts)}",
        f"- Gold causal claims per case: {min(claim_counts)}-{max(claim_counts)}",
    ]

    if coverage_report.issues:
        lines += ["", "## Coverage Issues", ""]
        for issue in coverage_report.issues:
            case_prefix = f"{issue.case_id}: " if issue.case_id else ""
            lines.append(f"- {issue.code}: {case_prefix}{issue.detail}")

    for case in dataset.cases:
        lines += [
            "",
            "---",
            "",
            f"## {case.case_id}: {case.case_title}",
            "",
            f"**Source PDF:** `{case.source_pdf}`",
            "",
            f"**Case Subject:** {case.case_subject}",
            "",
            "### Prompt",
            "",
            "```text",
            case.prompt,
            "```",
            "",
            "### Expected Output",
            "",
        ]
        for field in EXPECTED_OUTPUT_FIELDS:
            label = field.replace("_", " ").title()
            lines += [f"**{label}**", ""]
            lines.extend(f"- {bullet}" for bullet in getattr(case.expected_output, field))
            lines.append("")

        lines += ["### Categorization Targets", ""]
        for target in case.category_targets:
            lines.append(
                f"- {target.factor} -> {target.correct_category}: {target.rationale}"
            )

        lines += ["", "### Source Anchors", ""]
        lines.extend(f"- {anchor}" for anchor in case.validation_notes.source_anchors)

        lines += [
            "",
            "### Reread Justification",
            "",
            case.validation_notes.reread_justification,
        ]

        if case.avoid_claims:
            lines += ["", "### Avoid Claims", ""]
            lines.extend(f"- {claim}" for claim in case.avoid_claims)

    return "\n".join(lines).rstrip() + "\n"


def render_gold_judge_rubric_markdown(
    *,
    sample_case: GoldStandardCase | None = None,
    sample_response_text: str = "[MODEL RESPONSE TEXT GOES HERE]",
) -> str:
    """Render the strict judge prompt and structured-output contract."""
    lines = [
        "# Gold-Standard PESTEL Judge Rubric",
        "",
        "This is the exact system prompt and structured-output contract used by "
        "the LLM-as-a-judge pipeline for the 25-case PESTEL gold standard.",
        "",
        "## Fingerprints",
        "",
        f"- `judge_prompt_sha256`: `{gold_judge_prompt_sha256()}`",
        f"- `judge_schema_sha256`: `{gold_judge_schema_sha256()}`",
        "",
        "## System Prompt",
        "",
        "```text",
        GOLD_JUDGE_SYSTEM_PROMPT.rstrip(),
        "```",
        "",
        "## Structured Output Fields",
        "",
    ]

    schema = gold_judge_output_schema()
    properties = schema.get("properties", {})
    for field_name in [
        "response_claim_scores",
        "category_judgments",
        "missing_gold_claims",
        "unsupported_or_wrong_claims",
        "summary",
    ]:
        field_schema = properties.get(field_name, {})
        description = field_schema.get("description") or field_schema.get("title") or ""
        lines.append(f"- `{field_name}`: {description}".rstrip())

    lines += [
        "",
        "## Full Structured Output Schema",
        "",
        "```json",
        json.dumps(schema, indent=2, ensure_ascii=False),
        "```",
    ]

    if sample_case is not None:
        lines += [
            "",
            "## Sample Case-Specific User Prompt",
            "",
            f"Case: `{sample_case.case_id}`",
            "",
            "```text",
            build_gold_judge_user_prompt(sample_case, sample_response_text).rstrip(),
            "```",
        ]

    return "\n".join(lines).rstrip() + "\n"


def prompt_pack_csv_rows(dataset: GoldStandardDataset) -> list[dict[str, str]]:
    return [
        {
            "case_id": case.case_id,
            "case_title": case.case_title,
            "source_pdf": case.source_pdf,
            "prompt": case.prompt,
        }
        for case in dataset.cases
    ]


def gold_standard_manifest_payload(
    dataset: GoldStandardDataset,
    coverage_report: GoldStandardCoverageReport | None = None,
) -> dict[str, object]:
    """Return a compact freeze manifest for the gold-standard dataset."""
    coverage_report = coverage_report or validate_gold_standard_coverage(dataset)
    rows = _gold_standard_manifest_rows(dataset)
    return {
        "schema_version": dataset.schema_version,
        "dataset_sha256": model_sha256(dataset),
        "case_count": len(dataset.cases),
        "pdf_count": coverage_report.pdf_count,
        "coverage_valid": coverage_report.is_valid,
        "coverage_issue_count": len(coverage_report.issues),
        "cases": rows,
    }


def gold_standard_manifest_csv_rows(
    dataset: GoldStandardDataset,
) -> list[dict[str, str]]:
    """Return per-case freeze manifest rows suitable for CSV export."""
    return [
        {
            key: _csv_value(row[key])
            for key in GOLD_STANDARD_MANIFEST_CSV_FIELDS
        }
        for row in _gold_standard_manifest_rows(dataset)
    ]


def render_gold_standard_manifest_markdown(
    dataset: GoldStandardDataset,
    coverage_report: GoldStandardCoverageReport | None = None,
) -> str:
    """Render a compact case-level reproducibility manifest."""
    payload = gold_standard_manifest_payload(dataset, coverage_report)
    lines = [
        "# Gold-Standard Dataset Freeze Manifest",
        "",
        f"- Dataset SHA-256: `{payload['dataset_sha256']}`",
        f"- Cases: {payload['case_count']}",
        f"- PDF papers in inventory: {payload['pdf_count']}",
        f"- Dataset/PDF coverage valid: {_yes_no(bool(payload['coverage_valid']))}",
        f"- Coverage issues: {payload['coverage_issue_count']}",
        "",
        "| Case | Prompt SHA-256 | Expected Output SHA-256 | Claims | Targets | Anchors | Sections | Buckets |",
        "|---|---|---|---:|---:|---:|---|---|",
    ]
    for row in payload["cases"]:
        assert isinstance(row, dict)
        lines.append(
            "| "
            f"{row['case_id']} | "
            f"`{row['prompt_sha256']}` | "
            f"`{row['expected_output_sha256']}` | "
            f"{row['gold_claim_count']} | "
            f"{row['category_target_count']} | "
            f"{row['source_anchor_count']} | "
            f"{_yes_no(bool(row['expected_sections_complete']))} | "
            f"{_yes_no(bool(row['category_targets_cover_all_buckets']))} |"
        )
    return "\n".join(lines).rstrip() + "\n"


def strip_human_packet_sources(response_text: str) -> str:
    """Remove citation appendices from blinded human-review packet responses."""
    for marker in ("\n## Sources\n", "\n## Sources\r\n", "\n# Sources\n", "\n# Sources\r\n"):
        if marker in response_text:
            return response_text.split(marker, 1)[0].rstrip()
    return response_text


def render_human_rating_instructions() -> str:
    """Return a minimal rater tutorial for the human calibration workbook."""
    return "\n".join(
        [
            "# How to Rate",
            "",
            "**You get:** `rater_N_packet.md` (read) + `rater_N_ratings.xlsx` (fill, return).",
            "",
            "## Per case (you have 5)",
            "",
            "1. Read **Expected Output**, then **Response A** and **B** in the packet.",
            "2. In Excel, find rows with the same `case_id`.",
            "",
            "### Depth rows (`metric = analytical_depth`, 2 per case: A and B)",
            "- Fill **column H** with **1**, **2**, or **3** (use dropdown).",
            "- Leave **column I** empty.",
            "- **1** = surface · **2** = analytical · **3** = strategic (see packet rubric).",
            "",
            "### Category rows (`metric = categorization_accuracy`)",
            "- Fill **column I** with **true** or **false** (use dropdown).",
            "- Leave **column H** empty.",
            "- **true** = factor is discussed in the expected PESTEL bucket (`correct_category`).",
            "- **false** = missing, wrong bucket, or misleading.",
            "",
            "## Do not edit columns A–G",
            "",
            "Especially **not column B** (`rater_id` — already filled).",
            "",
            "Return your filled `rater_N_ratings.xlsx`.",
        ]
    ).rstrip() + "\n"


def render_human_packet_markdown(packet: HumanCalibrationPacket) -> str:
    lines = [
        "# Human Calibration Packet",
        "",
        packet.instructions,
        "",
        "## Rating Scale",
    ]
    for score, description in sorted(packet.rating_scale.items()):
        lines.append(f"- {score}: {description}")

    lines += [
        "",
        "## Ratings CSV Rules",
        "",
        "- Rows with `metric=analytical_depth`: fill `analytical_depth_score` with 1, 2, or 3; leave `correct` empty.",
        "- Rows with `metric=categorization_accuracy`: fill `correct` with true/false; leave `analytical_depth_score` empty.",
        "- Do not change `metric`, `rater_id`, `case_id`, `label`, `factor`, `correct_category`, or `rationale` values.",
        "- For category rows, use `factor`, `correct_category`, and `rationale` as the gold target and mark whether the response mapped that factor correctly.",
        "- Use the expected output as the reference standard, not as another model response.",
    ]

    for item in packet.items:
        lines += [
            "",
            "---",
            "",
            f"## {item.case_id}: {item.case_title}",
            "",
            "### Prompt",
            "",
            "```text",
            item.prompt,
            "```",
            "",
            "### Expected Output",
            "",
        ]
        for section, bullets in item.expected_output.items():
            lines.append(f"**{section.replace('_', ' ').title()}**")
            lines.extend(f"- {bullet}" for bullet in bullets)
            lines.append("")

        lines += ["### Categorization Targets", ""]
        for target in item.category_targets:
            lines.append(
                f"- {target['factor']} -> {target['correct_category']}: {target['rationale']}"
            )

        lines += ["", "### Anonymized Responses", ""]
        for output in item.outputs:
            lines += [
                f"#### Response {output.label}",
                "",
                "```text",
                strip_human_packet_sources(output.response_text),
                "```",
                "",
            ]
    return "\n".join(lines).rstrip() + "\n"


def ratings_template_csv_rows(template: HumanRatingsTemplate) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for rating in template.depth_ratings:
        rows.append(
            {
                "metric": "analytical_depth",
                "rater_id": rating.rater_id,
                "case_id": rating.case_id,
                "label": rating.label,
                "factor": "",
                "correct_category": "",
                "rationale": "",
                "analytical_depth_score": (
                    "" if rating.analytical_depth_score is None else str(rating.analytical_depth_score)
                ),
                "correct": "",
            }
        )
    for rating in template.category_ratings:
        rows.append(
            {
                "metric": "categorization_accuracy",
                "rater_id": rating.rater_id,
                "case_id": rating.case_id,
                "label": rating.label,
                "factor": rating.factor,
                "correct_category": rating.correct_category or "",
                "rationale": rating.rationale or "",
                "analytical_depth_score": "",
                "correct": "" if rating.correct is None else str(rating.correct).lower(),
            }
        )
    return rows


def _yes_no(value: bool) -> str:
    return "yes" if value else "no"


def _gold_standard_manifest_rows(
    dataset: GoldStandardDataset,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    expected_buckets = set(PESTEL_HEADINGS)
    for case in dataset.cases:
        expected_output = case.expected_output.model_dump(mode="json")
        gold_claims = [claim.model_dump(mode="json") for claim in case.gold_claims]
        category_targets = [
            target.model_dump(mode="json") for target in case.category_targets
        ]
        rows.append(
            {
                "case_id": case.case_id,
                "case_title": case.case_title,
                "source_pdf": case.source_pdf,
                "prompt_sha256": _text_sha256(case.prompt),
                "expected_output_sha256": _json_sha256(expected_output),
                "gold_claims_sha256": _json_sha256(gold_claims),
                "category_targets_sha256": _json_sha256(category_targets),
                "gold_claim_count": len(gold_claims),
                "category_target_count": len(category_targets),
                "source_anchor_count": len(case.validation_notes.source_anchors),
                "expected_sections_complete": all(
                    bool(getattr(case.expected_output, field))
                    for field in EXPECTED_OUTPUT_FIELDS
                ),
                "category_targets_cover_all_buckets": {
                    target.correct_category for target in case.category_targets
                }
                == expected_buckets,
            }
        )
    return rows


def _text_sha256(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _json_sha256(payload: object) -> str:
    canonical = json.dumps(payload, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def _csv_value(value: object) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def csv_text(rows: Sequence[dict[str, str]], fieldnames: Sequence[str]) -> str:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=fieldnames, lineterminator="\n")
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return buffer.getvalue()


def write_ratings_template_xlsx(
    path: str,
    rows: Sequence[dict[str, str]],
    fieldnames: Sequence[str] = RATINGS_CSV_FIELDS,
) -> None:
    """Write a rater workbook with dropdown validation on the editable columns."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Ratings"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    for column_index, field in enumerate(fieldnames, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=field)
        cell.font = header_font
        cell.fill = header_fill

    depth_col = list(fieldnames).index("analytical_depth_score") + 1
    correct_col = list(fieldnames).index("correct") + 1
    depth_validation = DataValidation(
        type="list",
        formula1='"1,2,3"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid depth score",
        error="Choose 1, 2, or 3 for analytical depth rows.",
    )
    correct_validation = DataValidation(
        type="list",
        formula1='"true,false"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid categorization answer",
        error="Choose true or false for categorization rows.",
    )

    for row_index, row in enumerate(rows, start=2):
        for column_index, field in enumerate(fieldnames, start=1):
            worksheet.cell(row=row_index, column=column_index, value=row.get(field, ""))
        if row.get("metric") == "analytical_depth":
            depth_validation.add(worksheet.cell(row=row_index, column=depth_col))
        elif row.get("metric") == "categorization_accuracy":
            correct_validation.add(worksheet.cell(row=row_index, column=correct_col))

    worksheet.add_data_validation(depth_validation)
    worksheet.add_data_validation(correct_validation)
    worksheet.freeze_panes = "A2"
    for column_index, field in enumerate(fieldnames, start=1):
        width = 28 if field in {"factor", "rationale"} else 18
        worksheet.column_dimensions[chr(64 + column_index)].width = width

    instructions = workbook.create_sheet("HowToRate")
    for line_index, line in enumerate(
        render_human_rating_instructions().splitlines(),
        start=1,
    ):
        instructions.cell(row=line_index, column=1, value=line)
    instructions.column_dimensions["A"].width = 100

    workbook.save(path)


def read_ratings_rows_from_file(path: str) -> list[dict[str, str]]:
    """Load rating rows from a CSV or XLSX rater file."""
    file_path = Path(path)
    if file_path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        worksheet = workbook["Ratings"] if "Ratings" in workbook.sheetnames else workbook.active
        rows_iter = worksheet.iter_rows(values_only=True)
        headers = [str(value) if value is not None else "" for value in next(rows_iter)]
        rows: list[dict[str, str]] = []
        for values in rows_iter:
            if not any(value not in (None, "") for value in values):
                continue
            row = {
                headers[index]: "" if value is None else str(value)
                for index, value in enumerate(values)
                if index < len(headers) and headers[index]
            }
            rows.append(row)
        return rows

    with file_path.open(newline="", encoding="utf-8") as handle:
        return [dict(row) for row in csv.DictReader(handle)]
